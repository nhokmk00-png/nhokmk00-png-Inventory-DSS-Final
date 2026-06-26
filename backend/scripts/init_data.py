from __future__ import annotations

import random
import sqlite3
import sys
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.app.config import DATA_DIR, settings

PRODUCTS = [
    ("SP001", "Áo thun nam cotton basic", "Áo thun", "Xưởng May An Phú", 129000, 76000, 4, 25, 12, "Chi nhánh trung tâm", 555, 5, 0, 65),
    ("SP002", "Quần jeans nữ ống rộng", "Quần jeans", "Denim House", 329000, 205000, 6, 18, 10, "Chi nhánh trung tâm", 370, 15, 0, 58),
    ("SP003", "Áo sơ mi nữ trắng công sở", "Áo sơ mi", "May Việt Style", 249000, 148000, 5, 20, 8, "Chi nhánh trung tâm", 26, 4, 0, 52),
    ("SP004", "Váy hoa nhí dáng xòe", "Váy nữ", "Lụa Mộc", 289000, 172000, 7, 16, 7, "Chi nhánh trung tâm", 6, 3, 0, 42),
    ("SP005", "Áo khoác dù nam chống nắng", "Áo khoác", "Outdoor Việt", 399000, 260000, 8, 12, 6, "Chi nhánh trung tâm", 280, 6, 0, 31),
    ("SP006", "Balo laptop chống sốc 15 inch", "Phụ kiện", "Túi Việt", 459000, 305000, 5, 14, 7, "Chi nhánh phụ 1", 82, 8, 12, 18),
    ("SP007", "Giày sneaker trắng basic", "Giày dép", "StepUp Factory", 499000, 330000, 6, 18, 9, "Chi nhánh phụ 2", 87, 12, 8, 24),
    ("SP008", "Túi tote canvas đi học", "Phụ kiện", "Canvas Home", 99000, 52000, 3, 35, 18, "Chi nhánh trung tâm", 100, 22, 5, 25),
    ("SP009", "Áo polo nam cổ bẻ", "Áo thun", "Xưởng May An Phú", 199000, 119000, 5, 22, 10, "Chi nhánh trung tâm", 38, 8, 15, 40),
    ("SP010", "Quần short kaki nam", "Quần short", "Kaki Việt", 189000, 108000, 4, 20, 10, "Chi nhánh trung tâm", 54, 10, 5, 34),
    ("SP011", "Chân váy chữ A màu đen", "Váy nữ", "May Việt Style", 219000, 128000, 4, 16, 8, "Chi nhánh phụ 1", 58, 7, 0, 22),
    ("SP012", "Áo croptop nữ tay ngắn", "Áo thun", "Lụa Mộc", 159000, 87000, 4, 18, 8, "Chi nhánh trung tâm", 328, 5, 0, 38),
    ("SP013", "Dép sandal nữ quai ngang", "Giày dép", "StepUp Factory", 259000, 160000, 6, 12, 6, "Chi nhánh phụ 2", 50, 6, 0, 16),
    ("SP014", "Mũ bucket kaki unisex", "Phụ kiện", "Kaki Việt", 89000, 41000, 3, 25, 12, "Chi nhánh phụ 1", 89, 13, 0, 15),
    ("SP015", "Áo len cardigan mỏng", "Áo khoác", "Len Việt", 349000, 218000, 8, 10, 5, "Chi nhánh trung tâm", 224, 3, 0, 24),
    ("SP016", "Đầm maxi đi biển", "Váy nữ", "Lụa Mộc", 379000, 225000, 7, 14, 6, "Chi nhánh phụ 2", 42, 4, 0, 55),
    ("SP017", "Áo hoodie unisex nỉ bông", "Áo khoác", "Outdoor Việt", 459000, 290000, 9, 12, 5, "Chi nhánh trung tâm", 18, 3, 0, 48),
    ("SP018", "Quần tây nam slimfit", "Quần dài", "Kaki Việt", 319000, 190000, 5, 15, 7, "Chi nhánh trung tâm", 64, 9, 0, 36),
    ("SP019", "Áo khoác bomber nữ", "Áo khoác", "Denim House", 429000, 270000, 8, 10, 5, "Chi nhánh phụ 1", 312, 4, 0, 28),
    ("SP020", "Túi đeo chéo mini nữ", "Phụ kiện", "Túi Việt", 189000, 98000, 4, 22, 10, "Chi nhánh trung tâm", 145, 12, 8, 62),
    ("SP021", "Giày loafer nữ da mềm", "Giày dép", "StepUp Factory", 569000, 380000, 7, 10, 5, "Chi nhánh phụ 2", 72, 6, 0, 18),
    ("SP022", "Áo dài cách tân nữ", "Áo dài", "May Việt Style", 689000, 430000, 10, 8, 4, "Chi nhánh trung tâm", 16, 2, 0, 50),
    ("SP023", "Set đồ thể thao nữ", "Đồ thể thao", "SportWear Việt", 349000, 205000, 6, 16, 8, "Chi nhánh phụ 1", 520, 10, 0, 30),
    ("SP024", "Kính mát thời trang", "Phụ kiện", "Canvas Home", 159000, 72000, 3, 30, 14, "Chi nhánh trung tâm", 90, 8, 0, 26),
    ("SP025", "Thắt lưng da nam", "Phụ kiện", "Kaki Việt", 229000, 120000, 4, 18, 8, "Chi nhánh phụ 2", 340, 5, 0, 20),
]

STATUS_PRIORITY = {"Nguy cấp": 1, "Cần nhập": 2, "Dư tồn": 3, "Bán chậm": 4, "Theo dõi": 5, "An toàn": 6}


def now_text() -> str:
    return datetime(2026, 6, 25, 9, 0, 0).isoformat(sep=" ")


def classify(stock: int, reserved: int, inbound: int, forecast_7d: int, lead: int, safety: int, minimum: int) -> dict:
    available = stock - reserved
    position = available + inbound
    avg = forecast_7d / 7
    reorder_point = avg * lead + safety
    target_stock = avg * (lead + 7) + safety
    suggested = max(0, round(target_stock - position))

    if stock > 300:
        return {
            "type": "OVERSTOCK", "status": "Dư tồn", "priority": STATUS_PRIORITY["Dư tồn"],
            "proposed": 0, "strategy": "Xem xét điều chuyển sang chi nhánh bán tốt, chạy khuyến mãi hoặc giảm nhập kỳ sau.",
            "reason": f"Tồn kho thực tế đang ở mức {stock}, cao hơn ngưỡng quản lý 300 sản phẩm.",
            "available": available, "position": position, "reorder": reorder_point, "target": target_stock,
        }
    if stock >= 180 and forecast_7d <= 40:
        return {
            "type": "SLOW_MOVING", "status": "Bán chậm", "priority": STATUS_PRIORITY["Bán chậm"],
            "proposed": 0, "strategy": "Tạm dừng nhập thêm, đề xuất khuyến mãi, combo hoặc điều chuyển hàng sang chi nhánh khác.",
            "reason": f"Tồn kho còn {stock} nhưng dự báo 7 ngày chỉ khoảng {forecast_7d} sản phẩm.",
            "available": available, "position": position, "reorder": reorder_point, "target": target_stock,
        }
    if position <= minimum:
        status = "Nguy cấp"; rec_type = "RESTOCK"; strategy = "Ưu tiên duyệt nhập ngay hoặc điều chuyển nội bộ trong ngày."
    elif position <= reorder_point:
        status = "Cần nhập"; rec_type = "RESTOCK"; strategy = "Lên kế hoạch nhập bổ sung theo số lượng đề xuất."
    elif position <= reorder_point * 1.3:
        status = "Theo dõi"; rec_type = "MONITOR"; strategy = "Theo dõi thêm nhu cầu bán và chuẩn bị phương án nhập nếu bán tăng."
    else:
        status = "An toàn"; rec_type = "NORMAL"; strategy = "Duy trì theo dõi định kỳ, chưa cần hành động ngay."
    return {
        "type": rec_type, "status": status, "priority": STATUS_PRIORITY[status], "proposed": suggested,
        "strategy": strategy, "reason": f"Vị thế tồn kho {position:.0f}, điểm đặt hàng lại {reorder_point:.0f}, dự báo 7 ngày {forecast_7d}.",
        "available": available, "position": position, "reorder": reorder_point, "target": target_stock,
    }


def create_template() -> None:
    path = DATA_DIR / "templates" / "Mau_Nhap_Xuat_Kho.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "NhapXuatKho"
    headers = ["Ngày chứng từ", "Loại phiếu", "Mã sản phẩm", "Tên sản phẩm", "Số lượng", "Đơn giá thực tế", "Người nhập", "Ghi chú"]
    ws.append(headers)
    rows = [
        ["2026-06-25", "Phiếu nhập hàng", "SP003", "Áo sơ mi nữ trắng công sở", 20, 151000, "Nhân viên kho", "Nhập bổ sung hàng bán chạy"],
        ["2026-06-25", "Phiếu xuất bán", "SP001", "Áo thun nam cotton basic", 35, 118000, "Nhân viên kho", "Xuất ưu đãi cho khách hàng lớn"],
        ["2026-06-25", "Phiếu điều chỉnh kiểm kê", "SP004", "Váy hoa nhí dáng xòe", 5, 172000, "Nhân viên kho", "Kiểm kê bổ sung"],
    ]
    for row in rows:
        ws.append(row)
    fill = PatternFill(fill_type="solid", fgColor="17365D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
    widths = [16, 28, 14, 36, 12, 18, 22, 42]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    wb.save(path)

    sample = ROOT / "import_samples" / "Phieu_Nhap_Xuat_Mau.xlsx"
    sample.parent.mkdir(parents=True, exist_ok=True)
    wb.save(sample)


def init_database() -> None:
    random.seed(20260625)
    settings.database_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(settings.database_path)
    connection.row_factory = sqlite3.Row
    schema = (ROOT / "backend" / "schema.sql").read_text(encoding="utf-8")
    connection.executescript(schema)

    connection.executemany(
        """
        INSERT INTO products (product_id, product_name, category, supplier_name, unit_price, purchase_price,
          lead_time_days, safety_stock, minimum_stock, branch_name)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [(p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7], p[8], p[9]) for p in PRODUCTS],
    )
    connection.executemany(
        "INSERT INTO inventory VALUES (?, ?, ?, ?, ?)",
        [(p[0], p[10], p[11], p[12], now_text()) for p in PRODUCTS],
    )

    for p in PRODUCTS:
        pid, forecast_7d = p[0], p[13]
        base = max(1, int(forecast_7d / 7))
        for i in range(60):
            day = date(2026, 4, 27) + timedelta(days=i)
            promo = 1 if i in [6, 7, 14] and pid in {"SP001", "SP002", "SP003"} else 0
            demand = max(0, base + (5 if promo else 0) + random.randint(-2, 2))
            if pid in {"SP001", "SP002", "SP012", "SP015"}:
                demand = max(0, int(demand * 0.45))
            connection.execute("INSERT INTO daily_demand VALUES (?, ?, ?, ?, ?)", (f"DD_{pid}_{day}", day.isoformat(), pid, demand, promo))
        forecast_id = f"FC_BATCH_20260625_001_{pid}"
        connection.execute(
            "INSERT INTO forecasts VALUES (?, 'BATCH_20260625_001', ?, ?, ?, ?, ?, ?, ?, ?)",
            (forecast_id, "2026-06-25", pid, float(forecast_7d), float(forecast_7d) / 7, "Weighted Moving Average", round(random.uniform(1.1, 5.8), 2), round(random.uniform(0.08, 0.25), 3), now_text()),
        )

    for row in connection.execute(
        """
        SELECT p.*, i.stock_on_hand, i.reserved_quantity, i.confirmed_inbound_quantity,
               f.forecast_id, f.forecast_quantity_7_days
        FROM products p JOIN inventory i ON i.product_id=p.product_id
        JOIN forecasts f ON f.product_id=p.product_id
        """
    ).fetchall():
        data = classify(row["stock_on_hand"], row["reserved_quantity"], row["confirmed_inbound_quantity"], row["forecast_quantity_7_days"], row["lead_time_days"], row["safety_stock"], row["minimum_stock"])
        rec_id = f"REC_{row['product_id']}_20260625"
        connection.execute(
            """
            INSERT INTO recommendations VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Chờ xử lý', NULL, NULL, NULL, 'Chưa gửi', 'Không áp dụng', NULL, NULL, NULL, ?)
            """,
            (rec_id, row["forecast_id"], row["product_id"], data["type"], data["status"], data["priority"], data["available"], data["position"], data["reorder"], data["target"], data["proposed"], None, data["strategy"], data["reason"], now_text()),
        )
        if data["status"] != "An toàn":
            connection.execute(
                "INSERT INTO ai_insights VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    f"AI_{row['product_id']}_20260625", rec_id,
                    f"{row['product_name']} đang ở trạng thái {data['status']}.",
                    data["reason"], data["strategy"],
                    "Thông tin này hỗ trợ quản lý xem xét phương án xử lý phù hợp.",
                    "FALLBACK", "Gemini chưa cấu hình, dùng nội dung gợi ý nội bộ.", now_text(),
                ),
            )

    tx_rows = [
        ("TX001", "SP001", "RECEIPT", 35, 520, 555, "Phiếu nhập hàng", 78000, 2730000, "Giá nhập NCC tăng nhẹ", "Nhập hàng sau livestream", "Nhân viên kho Minh", "2026-06-21 08:30:00"),
        ("TX002", "SP002", "SALE_OUT", 18, 388, 370, "Phiếu xuất bán", 309000, 5562000, "Giá ưu đãi cho khách hàng thân thiết", "Xuất bán tại cửa hàng", "Nhân viên bán hàng", "2026-06-22 10:10:00"),
        ("TX003", "SP004", "ADJUSTMENT_OUT", 6, 12, 6, "Phiếu hàng lỗi / hủy", 172000, 1032000, "Tính theo giá vốn", "Hàng lỗi sau kiểm kê", "Nhân viên kho Hạnh", "2026-06-23 17:40:00"),
        ("TX004", "SP003", "RECEIPT", 20, 6, 26, "Phiếu nhập hàng", 151000, 3020000, "Giá nhập mới từ NCC", "Nhập bổ sung hàng công sở", "Nhân viên kho Minh", "2026-06-24 09:15:00"),
        ("TX005", "SP005", "TRANSFER_OUT", 15, 295, 280, "Phiếu điều chuyển", 260000, 3900000, "Theo giá vốn chuyển kho", "Điều chuyển sang chi nhánh phụ 2", "Quản lý kho", "2026-06-24 14:30:00"),
        ("TX006", "SP020", "SALE_OUT", 24, 145, 121, "Phiếu xuất bán", 175000, 4200000, "Giá combo phụ kiện", "Khách đoàn mua phụ kiện", "Nhân viên bán hàng", "2026-06-25 10:20:00"),
        ("TX007", "SP017", "RECEIPT", 40, 18, 58, "Phiếu nhập hàng", 288000, 11520000, "Nhập trước mùa lạnh", "Bổ sung hoodie bán chạy", "Nhân viên kho Minh", "2026-06-25 11:10:00"),
        ("TX008", "SP023", "TRANSFER_OUT", 60, 520, 460, "Phiếu điều chuyển", 205000, 12300000, "Giảm dư tồn ở chi nhánh chính", "Điều chuyển sang chi nhánh bán tốt", "Quản lý kho", "2026-06-25 15:35:00"),
        ("TX009", "SP022", "RECEIPT", 25, 16, 41, "Phiếu nhập hàng", 435000, 10875000, "Giá nhập đợt mới", "Chuẩn bị dịp lễ", "Nhân viên kho Hạnh", "2026-06-26 09:05:00"),
        ("TX010", "SP001", "SALE_OUT", 50, 555, 505, "Phiếu xuất bán", 115000, 5750000, "Giá ưu đãi khách hàng lớn", "Bán theo đơn khách sỉ", "Nhân viên bán hàng", "2026-06-26 14:30:00"),
    ]
    connection.executemany("INSERT INTO inventory_transactions VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", tx_rows)

    imports = [
        ("IMP001", "2026-06-21 08:20:00", "phieu_nhap_21_06.xlsx", "Nhân viên kho Minh", 18, 0, "Đã nhập", "Nhập hàng bổ sung"),
        ("IMP002", "2026-06-22 17:35:00", "kiem_ke_cuoi_ngay.xlsx", "Nhân viên kho Hạnh", 12, 1, "Có lỗi", "Một dòng sai mã sản phẩm"),
        ("IMP003", "2026-06-23 08:45:00", "xuat_ban_khach_si.xlsx", "Nhân viên bán hàng", 9, 0, "Đã nhập", "Xuất bán theo đơn khách sỉ"),
        ("IMP004", "2026-06-24 14:30:00", "dieu_chuyen_chi_nhanh.xlsx", "Quản lý kho", 11, 0, "Đã nhập", "Điều chuyển hàng tồn cao"),
        ("IMP005", "2026-06-25 09:05:00", "de_xuat_nhap_sang.xlsx", "Quản lý cửa hàng", 15, 0, "Đã nhập", "Dữ liệu vận hành đầu ngày"),
        ("IMP006", "2026-06-26 09:15:00", "cap_nhat_gia_nhap_xuat.xlsx", "Nhân viên kho Hạnh", 10, 0, "Đã nhập", "Cập nhật phiếu có đơn giá thực tế"),
    ]
    connection.executemany("INSERT INTO excel_imports VALUES (?, ?, ?, ?, ?, ?, ?, ?)", imports)

    connection.commit()
    errors = connection.execute("PRAGMA foreign_key_check").fetchall()
    connection.close()
    if errors:
        raise RuntimeError(errors)
    create_template()
    print(f"Đã khởi tạo dữ liệu vận hành: {settings.database_path}")


if __name__ == "__main__":
    init_database()
