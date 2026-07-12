from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import ExcelImport, ExcelImportRow, Product
from .inventory_service import create_inventory_transaction

HEADERS = ["Ngày chứng từ", "Loại phiếu", "Mã sản phẩm", "Tên sản phẩm", "Số lượng", "Đơn giá thực tế", "Người nhập", "Ghi chú"]
VALID_VOUCHERS = ["Phiếu nhập hàng", "Phiếu xuất bán", "Phiếu điều chỉnh kiểm kê", "Phiếu hàng lỗi / hủy", "Phiếu điều chuyển chi nhánh"]


def make_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "phieu_kho"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")
    samples = [
        ["2026-07-07", "Phiếu nhập hàng", "SP004", "Váy hoa nhí dáng xòe", 40, 89000, "Nhân viên kho", "Nhập bổ sung theo đề xuất"],
        ["2026-07-07", "Phiếu xuất bán", "SP001", "Áo thun nam cotton basic", 10, 115000, "Nhân viên bán hàng", "Xuất ưu đãi cho khách hàng lớn"],
        ["2026-07-07", "Phiếu điều chuyển chi nhánh", "SP002", "Quần jeans nữ ống rộng", 30, 145000, "Quản lý kho", "Điều chuyển sang chi nhánh thiếu hàng"],
    ]
    for row in samples:
        ws.append(row)
    for idx, width in enumerate([16, 28, 14, 34, 12, 18, 22, 42], start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _normalize_row(row) -> dict:
    values = list(row)
    while len(values) < len(HEADERS):
        values.append(None)
    return dict(zip(HEADERS, values))


def upload_workbook(db: Session, file_bytes: bytes, file_name: str, uploaded_by: str) -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [h for h in HEADERS if h not in header]
    import_file = ExcelImport(file_name=file_name, uploaded_by=uploaded_by)
    db.add(import_file)
    db.flush()

    if missing:
        import_file.error_rows = max(1, ws.max_row - 1)
        import_file.status = "Có lỗi"
        db.add(ExcelImportRow(import_id=import_file.import_id, row_number=1, row_status="Lỗi", error_message=f"Thiếu cột: {', '.join(missing)}"))
        db.commit()
        return import_dict(import_file, include_rows=True)

    valid_rows = 0
    error_rows = 0
    header_map = {name: header.index(name) for name in HEADERS}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {name: row[header_map[name]] for name in HEADERS}
        product_id = str(data["Mã sản phẩm"] or "").strip()
        voucher_type = str(data["Loại phiếu"] or "").strip()
        quantity = int(data["Số lượng"] or 0) if str(data["Số lượng"] or "").strip() else 0
        unit_price = float(data["Đơn giá thực tế"] or 0)
        product = db.get(Product, product_id)
        errors = []
        if not product:
            errors.append("Mã sản phẩm không tồn tại")
        if voucher_type not in VALID_VOUCHERS:
            errors.append("Loại phiếu không hợp lệ")
        if quantity <= 0:
            errors.append("Số lượng phải lớn hơn 0")
        if unit_price < 0:
            errors.append("Đơn giá không được âm")
        current_stock = product.inventory.stock_on_hand if product and product.inventory else (product.stock_on_hand if product else 0)
        if product and voucher_type != "Phiếu nhập hàng" and current_stock < quantity:
            errors.append("Tồn kho không đủ để xuất/điều chuyển/hủy")
        row_obj = ExcelImportRow(
            import_id=import_file.import_id,
            row_number=row_idx,
            document_date=str(data["Ngày chứng từ"] or ""),
            voucher_type=voucher_type,
            product_id=product_id,
            product_name=str(data["Tên sản phẩm"] or (product.product_name if product else "")),
            quantity=quantity,
            unit_price=unit_price,
            uploaded_by=str(data["Người nhập"] or uploaded_by),
            note=str(data["Ghi chú"] or ""),
            row_status="Lỗi" if errors else "Hợp lệ",
            error_message="; ".join(errors),
        )
        db.add(row_obj)
        if errors:
            error_rows += 1
            continue
        direction = "IN" if voucher_type == "Phiếu nhập hàng" else "OUT"
        if voucher_type == "Phiếu điều chỉnh kiểm kê":
            direction = "IN" if "tăng" in row_obj.note.lower() else "OUT"
        try:
            create_inventory_transaction(db, {
                "product_id": product_id,
                "voucher_type": voucher_type,
                "movement_direction": direction,
                "quantity": quantity,
                "unit_price": unit_price,
                "price_note": "Theo đơn giá thực tế trong file Excel",
                "recorded_by": row_obj.uploaded_by,
                "reason": row_obj.note or f"Import từ file {file_name}",
            })
            valid_rows += 1
        except ValueError as exc:
            row_obj.row_status = "Lỗi"
            row_obj.error_message = str(exc)
            error_rows += 1
    import_file.valid_rows = valid_rows
    import_file.error_rows = error_rows
    import_file.status = "Đã nhập" if error_rows == 0 else "Có lỗi"
    db.commit()
    db.refresh(import_file)
    return import_dict(import_file, include_rows=True)


def import_dict(import_file: ExcelImport, include_rows: bool = False) -> dict:
    data = {
        "import_id": import_file.import_id,
        "import_date": import_file.import_date,
        "file_name": import_file.file_name,
        "uploaded_by": import_file.uploaded_by,
        "valid_rows": import_file.valid_rows,
        "error_rows": import_file.error_rows,
        "status": import_file.status,
    }
    if include_rows:
        data["rows"] = [row_dict(r) for r in import_file.rows]
    return data


def row_dict(row: ExcelImportRow) -> dict:
    return {
        "row_id": row.row_id,
        "row_number": row.row_number,
        "document_date": row.document_date,
        "voucher_type": row.voucher_type,
        "product_id": row.product_id,
        "product_name": row.product_name,
        "quantity": row.quantity,
        "unit_price": row.unit_price,
        "uploaded_by": row.uploaded_by,
        "note": row.note,
        "row_status": row.row_status,
        "error_message": row.error_message,
    }
