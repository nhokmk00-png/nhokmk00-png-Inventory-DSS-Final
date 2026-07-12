from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Product, Recommendation
from .inventory_service import recommendation_dict


def build_recommendations_workbook(db: Session) -> bytes:
    """Tạo báo cáo Excel xử lý tồn kho để tải xuống hoặc gửi qua Telegram."""
    wb = Workbook()
    ws = wb.active
    ws.title = "xu_ly_ton_kho"

    headers = [
        "Mã SP",
        "Tên sản phẩm",
        "Tình trạng",
        "Tồn",
        "ROP",
        "Dự báo 7 ngày",
        "Đề xuất nhập ban đầu",
        "Điều chuyển ban đầu",
        "Số lượng quản lý chốt",
        "Xử lý nội bộ",
        "Thông báo",
        "Phương án thực hiện",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = db.scalars(
        select(Recommendation)
        .join(Product)
        .where(Product.is_active == 1)
        .order_by(Recommendation.recommendation_id)
    ).all()
    for rec in rows:
        data = recommendation_dict(rec)
        ws.append(
            [
                data["product_id"],
                data["product_name"],
                data["business_status"],
                data["stock_on_hand"],
                data["reorder_point"],
                data["forecast_7_days"],
                data["proposed_quantity"],
                data["transfer_quantity"],
                data["final_quantity"],
                data["internal_status"],
                data["telegram_status"],
                data["effective_action_strategy"],
            ]
        )

    widths = [12, 32, 16, 10, 10, 14, 20, 20, 22, 18, 18, 62]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
