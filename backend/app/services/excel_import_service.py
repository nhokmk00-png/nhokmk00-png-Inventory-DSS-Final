from __future__ import annotations

import io
import uuid
from datetime import datetime
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook

from backend.app.config import DATA_DIR
from backend.app.database import connect, transaction
from backend.app.services.inventory_service import create_inventory_voucher, resolve_product_id

TEMPLATE_PATH = DATA_DIR / "templates" / "Mau_Nhap_Xuat_Kho.xlsx"
EXPECTED_HEADERS = ["Ngày chứng từ", "Loại phiếu", "Mã sản phẩm", "Tên sản phẩm", "Số lượng", "Đơn giá thực tế", "Người nhập", "Ghi chú"]
VOUCHER_OPTIONS = {"Phiếu nhập hàng", "Phiếu xuất bán", "Phiếu điều chỉnh kiểm kê", "Phiếu hàng lỗi / hủy", "Phiếu điều chuyển chi nhánh"}


def now_text() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def import_history(limit: int = 50) -> list[dict]:
    with connect() as connection:
        return [dict(row) for row in connection.execute("SELECT * FROM excel_imports ORDER BY import_date DESC LIMIT ?", (limit,)).fetchall()]


def _cell(row, index: int):
    return row[index] if index < len(row) else None


async def process_upload(file: UploadFile, uploaded_by: str) -> dict:
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb["NhapXuatKho"] if "NhapXuatKho" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File Excel trống.")
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    missing = [h for h in EXPECTED_HEADERS if h not in headers]
    if missing:
        raise ValueError("File thiếu cột: " + ", ".join(missing))
    idx = {h: headers.index(h) for h in EXPECTED_HEADERS}
    import_id = f"IMP_{uuid.uuid4().hex[:10].upper()}"
    details = []
    valid_rows = 0
    error_rows = 0
    with connect() as connection:
        for number, values in enumerate(rows[1:], start=2):
            if not any(values):
                continue
            voucher_type = str(_cell(values, idx["Loại phiếu"]) or "").strip()
            product_raw = str(_cell(values, idx["Mã sản phẩm"]) or _cell(values, idx["Tên sản phẩm"]) or "").strip()
            product_name = str(_cell(values, idx["Tên sản phẩm"]) or "").strip()
            qty_raw = _cell(values, idx["Số lượng"])
            unit_price_raw = _cell(values, idx["Đơn giá thực tế"])
            recorded_by = str(_cell(values, idx["Người nhập"]) or uploaded_by or "Nhân viên kho").strip()
            note = str(_cell(values, idx["Ghi chú"]) or "Nhập file Excel").strip()
            status = "Hợp lệ"; error = ""; quantity = 0; product_id = None; unit_price = None
            try:
                quantity = int(qty_raw)
                if quantity <= 0:
                    raise ValueError("Số lượng phải lớn hơn 0")
                if voucher_type not in VOUCHER_OPTIONS:
                    raise ValueError("Loại phiếu không hợp lệ")
                if unit_price_raw not in (None, ""):
                    unit_price = float(str(unit_price_raw).replace(".", "").replace(",", ""))
                    if unit_price < 0:
                        raise ValueError("Đơn giá không được âm")
                product_id = resolve_product_id(connection, product_raw or product_name)
                if not product_id:
                    raise ValueError("Không tìm thấy sản phẩm")
            except Exception as exc:
                status = "Lỗi"; error = str(exc); error_rows += 1
            else:
                valid_rows += 1
            details.append({"row_id": f"ROW_{uuid.uuid4().hex[:12].upper()}", "import_id": import_id, "excel_row": number, "product_id": product_id, "product_name": product_name, "voucher_type": voucher_type, "quantity": quantity, "unit_price": unit_price, "recorded_by": recorded_by, "note": note, "row_status": status, "error_message": error})
    for row in [r for r in details if r["row_status"] == "Hợp lệ"]:
        create_inventory_voucher(row["product_id"], row["voucher_type"], row["quantity"], row["note"], row["recorded_by"], unit_price=row.get("unit_price"), price_note="Đơn giá từ file Excel" if row.get("unit_price") is not None else None)
    status = "Đã nhập" if error_rows == 0 else "Có lỗi"
    with transaction() as connection:
        connection.execute("INSERT INTO excel_imports VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (import_id, now_text(), file.filename or "upload.xlsx", uploaded_by, valid_rows, error_rows, status, "Đã xử lý file Excel"))
        connection.executemany("INSERT INTO excel_import_rows VALUES (?, ?, ?, ?, ?, ?, ?, ?)", [(d["row_id"], import_id, d["product_id"], d["product_name"], d["voucher_type"], d["quantity"], d["row_status"], d["error_message"]) for d in details])
    return {"import_id": import_id, "file_name": file.filename, "uploaded_by": uploaded_by, "valid_rows": valid_rows, "error_rows": error_rows, "status": status, "rows": details}
